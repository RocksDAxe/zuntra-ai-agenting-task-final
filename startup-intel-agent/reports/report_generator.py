"""
ReportGenerator
---------------
Builds the Executive Report from the structured dataset produced by the
Orchestrator, and exports it as Markdown, PDF, and an Excel workbook
(bonus deliverable: PDF/Excel export).
"""
from datetime import datetime
from typing import Dict

import config


class ReportGenerator:
    def __init__(self, result: Dict):
        self.result = result
        self.startups = sorted(
            result["startups"],
            key=lambda s: s["insights"]["momentum_score"],
            reverse=True,
        )

    # ------------------------------------------------------------------ #
    # Markdown
    # ------------------------------------------------------------------ #
    def build_markdown(self) -> str:
        lines = []
        lines.append("# AI Startup Intelligence — Executive Report")
        lines.append(f"_Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}_\n")

        lines.append("## Portfolio Summary")
        lines.append(f"- Startups tracked: **{len(self.startups)}**")
        lines.append(f"- Raw articles ingested: **{self.result['raw_article_count']}**")
        lines.append(f"- Signals before de-duplication: **{self.result['signal_count_before_dedup']}**")
        lines.append(f"- Unique signals after consolidation: **{self.result['signal_count_after_dedup']}**")
        dupes_removed = self.result['signal_count_before_dedup'] - self.result['signal_count_after_dedup']
        lines.append(f"- Duplicate reports merged: **{dupes_removed}**\n")

        strong_buys = [s for s in self.startups if s["insights"]["recommendation"] == "Strong Buy Signal"]
        watch = [s for s in self.startups if s["insights"]["recommendation"] == "Watch — Emerging Momentum"]
        caution = [s for s in self.startups if "Caution" in s["insights"]["recommendation"]]

        lines.append("## Top Investment Signals")
        if strong_buys:
            lines.append("**Strong Buy Signal:** " + ", ".join(s["name"] for s in strong_buys))
        if watch:
            lines.append("**Watch:** " + ", ".join(s["name"] for s in watch))
        if caution:
            lines.append("**Caution / Elevated Risk:** " + ", ".join(s["name"] for s in caution))
        lines.append("")

        lines.append("## Startup Profiles\n")
        for s in self.startups:
            lines.append(f"### {s['name']} ({s['sector']})")
            lines.append(f"- Headquarters: {s['headquarters']} | Founded: {s['founded_year']} | "
                         f"Est. employees: {s['employee_estimate']}")
            lines.append(f"- **Predicted stage:** {s['stage']['predicted_stage']} "
                         f"(confidence {s['stage']['confidence']:.0%})")
            lines.append(f"- **Total funding tracked:** ${s['stage']['total_funding_musd']:.1f}M "
                         f"| Latest round: {s['stage']['latest_funding_stage']}")
            lines.append(f"- **Momentum score:** {s['insights']['momentum_score']}/100 | "
                         f"**Risk score:** {s['insights']['risk_score']}/100 | "
                         f"**Recommendation:** {s['insights']['recommendation']}")
            lines.append(f"- {s['insights']['rationale']}")
            lines.append("- Opportunities:")
            for o in s["insights"]["opportunities"]:
                lines.append(f"  - {o}")
            lines.append("- Risks:")
            for r in s["insights"]["risks"]:
                lines.append(f"  - {r}")
            lines.append("")

        return "\n".join(lines)

    def save_markdown(self):
        md = self.build_markdown()
        with open(config.REPORT_MD, "w", encoding="utf-8") as f:
            f.write(md)
        print(f"[ReportGenerator] Saved markdown report -> {config.REPORT_MD}")
        return md

    # ------------------------------------------------------------------ #
    # PDF (uses fpdf2; pure-python, no system deps)
    # ------------------------------------------------------------------ #
    def save_pdf(self):
        try:
            from fpdf import FPDF
        except ImportError:
            print("[ReportGenerator] fpdf2 not installed; skipping PDF export. "
                  "Install with: pip install fpdf2")
            return

        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()

        def cell(text, size=10, bold=False, line_h=6, gap=0):
            # fpdf2 leaves the X cursor at the right margin after a full-width
            # multi_cell; reset it before every call or width collapses to 0.
            pdf.set_x(pdf.l_margin)
            pdf.set_font("Helvetica", "B" if bold else "", size)
            pdf.multi_cell(0, line_h, text.encode("latin-1", "replace").decode("latin-1"))
            if gap:
                pdf.ln(gap)

        cell("AI Startup Intelligence - Executive Report", size=16, bold=True, line_h=10)
        cell(f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", gap=4)

        cell("Portfolio Summary", size=12, bold=True, line_h=8)
        cell(
            f"Startups tracked: {len(self.startups)}\n"
            f"Raw articles ingested: {self.result['raw_article_count']}\n"
            f"Signals before dedup: {self.result['signal_count_before_dedup']}\n"
            f"Unique signals after dedup: {self.result['signal_count_after_dedup']}",
            gap=4,
        )

        for s in self.startups:
            cell(f"{s['name']} ({s['sector']})", size=12, bold=True, line_h=8)
            body = (
                f"HQ: {s['headquarters']} | Founded: {s['founded_year']} | "
                f"Employees (est): {s['employee_estimate']}\n"
                f"Predicted stage: {s['stage']['predicted_stage']} "
                f"(confidence {s['stage']['confidence']:.0%})\n"
                f"Total funding tracked: ${s['stage']['total_funding_musd']:.1f}M | "
                f"Latest round: {s['stage']['latest_funding_stage']}\n"
                f"Momentum: {s['insights']['momentum_score']}/100  "
                f"Risk: {s['insights']['risk_score']}/100  "
                f"Recommendation: {s['insights']['recommendation']}\n"
                f"{s['insights']['rationale']}\n"
                "Opportunities:\n" + "\n".join(f" - {o}" for o in s["insights"]["opportunities"]) +
                "\nRisks:\n" + "\n".join(f" - {r}" for r in s["insights"]["risks"])
            )
            cell(body, size=10, line_h=6, gap=3)

        pdf.output(str(config.REPORT_PDF))
        print(f"[ReportGenerator] Saved PDF report -> {config.REPORT_PDF}")

    # ------------------------------------------------------------------ #
    # Excel workbook (uses openpyxl)
    # ------------------------------------------------------------------ #
    def save_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            print("[ReportGenerator] openpyxl not installed; skipping Excel export. "
                  "Install with: pip install openpyxl")
            return

        wb = Workbook()

        # -- Sheet 1: Startup summary --
        ws = wb.active
        ws.title = "Startup Summary"
        headers = ["Name", "Sector", "HQ", "Founded", "Employees (est)", "Predicted Stage",
                   "Confidence", "Total Funding ($M)", "Latest Round", "Momentum",
                   "Risk", "Recommendation", "Signal Count"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F2937")

        for s in self.startups:
            ws.append([
                s["name"], s["sector"], s["headquarters"], s["founded_year"],
                s["employee_estimate"], s["stage"]["predicted_stage"],
                round(s["stage"]["confidence"], 2), s["stage"]["total_funding_musd"],
                s["stage"]["latest_funding_stage"], s["insights"]["momentum_score"],
                s["insights"]["risk_score"], s["insights"]["recommendation"],
                s["signal_count"],
            ])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 16

        # -- Sheet 2: Signals detail --
        ws2 = wb.create_sheet("Signals Detail")
        ws2.append(["Startup", "Signal Type", "Date", "Title", "Sources", "Source Count"])
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F2937")
        for s in self.startups:
            for sig in s["signals"]:
                ws2.append([s["name"], sig["signal_type"], sig["date"], sig["title"],
                            "; ".join(sig.get("sources", [])), sig.get("source_count", 1)])
        for col in ws2.columns:
            ws2.column_dimensions[col[0].column_letter].width = 22

        wb.save(config.REPORT_XLSX)
        print(f"[ReportGenerator] Saved Excel workbook -> {config.REPORT_XLSX}")

    def generate_all(self):
        self.save_markdown()
        self.save_pdf()
        self.save_excel()
